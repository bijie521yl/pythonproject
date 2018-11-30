# coding=GBK
# tc40��־
from time import sleep
from past.translation import common_substring
from encodings import undefined
__author__ = 'bj'
# import httplib
# import urllib

import urllib2
import json
import datetime
import time

import ConfigParser
import threading



from openpyxl import load_workbook

#test big data for China Securities Co., Ltd.

from openpyxl import Workbook
import os

cf = ConfigParser.ConfigParser() 
cf.read("fileset.ini")

g_filepath=cf.get("file", "filepath");
g_apitype=cf.getint("file", "apitype");
g_zjzh=cf.get("file", "zjzh");
g_excelfile=cf.get("file", "excelfile");


def RriteExcel(filepath,listexcel):
    if os.path.exists(filepath):
        wb = load_workbook(filename=filepath)
        sheet_names = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheet_names[0])
        for index in range(len(listexcel)):
            if index==0:
                continue;
            sheet.append(listexcel[index])
        wb.save(filepath)
    else:
        
        wb = Workbook()
        sheet_names = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheet_names[0])
        for index in range(len(listexcel)):
            sheet.append(listexcel[index])
        wb.save(filepath)
#     wb.save(r'D:\360Downloads\example.xlsx')

def jwkj_get_filePath_fileName_fileExt(filename):  
    (filepath,tempfilename) = os.path.split(filename);  
    (shotname,extension) = os.path.splitext(tempfilename);  
#     print filepath,shotname,extension
    return [filepath,shotname,extension]  


def ReadLog(filepath,file,hzm,apitype,zjzh):
    #ָ��·����Э���
    os.chdir(filepath)
    fdata=open(file+hzm)
    LineTemp = fdata.readline()
    listexcel=[]
    if (apitype==100):
        findstr="Э��:100"
    elif(apitype==101):
        findstr="Э��:101"
    else:
        
        findstr="Э��:202"
    while LineTemp:
        linelist=[]
        if LineTemp.find(findstr) >= 0: 
            if len(LineTemp.strip())>0:
                LineTemp=LineTemp[0:-2] 
                 
            list1=LineTemp.split(" ");
            if len(list1)>0:
                if findstr=="Э��:100":
                    linelist.append(file);
                    for index in range(len(list1)):
                        temp=list1[index].split(":")
                        if len(temp)==2:
                            linelist.append(temp[1])
                    LineTemp = fdata.readline()
                    if len(LineTemp.strip())>0:
                        LineTemp=LineTemp[0:-2] 
                    LineTemp1=LineTemp.split("|");
                    linelist.extend(LineTemp1)
#                     if zjzh==undefined:
#                         if len(LineTemp1)>1:
#                             linelist.extend(LineTemp1)
#                         if len(linelist)!=0:
#                             listexcel.append(linelist)  
#                     else:
#                         if LineTemp1[0]==zjzh:    
#                             if len(LineTemp1)>1:
#                                 linelist.extend(LineTemp1)
#                             if len(linelist)!=0:
#                                 listexcel.append(linelist)
#                     LineTemp = fdata.readline()
                    #�������϶���
                    
                    LineTemp = fdata.readline()
#                     for index in range(len(list1)):
#                         temp=list1[index].split(":")
#                         if len(temp)==2:
#                             linelist.append(temp[1])
                            
                    LineTemp = fdata.readline()
                    if len(LineTemp.strip())>0:
                        LineTemp=LineTemp[0:-2] 
                    LineTemp1=LineTemp.split("|");
                    if LineTemp1[0]=="0" and LineTemp1[2]=="1":
                        LineTemp = fdata.readline()
                        if len(LineTemp.strip())>0:
                            LineTemp=LineTemp[0:-2] 
                        LineTemp1=LineTemp.split("|");
                    
                        if zjzh==undefined:
                            if len(LineTemp1)>1:
                                
                                linelist.extend([LineTemp1[0]])
                            if len(linelist)!=0:
                                listexcel.append([LineTemp1[0]])  
                        else:
                            if LineTemp1[0]==zjzh:    
#                                 print LineTemp1[1]
                                if len(LineTemp1)>1:
                                    linelist.extend([LineTemp1[0]])
                                if len(linelist)!=0:
                                    listexcel.append(linelist)
                    LineTemp = fdata.readline()
                    
                    
#                 elif findstr=="Э��:101":
#                     linelist.append(file);
#                     for index in range(len(list1)):
#                         temp=list1[index].split(":")
#                         if len(temp)==2:
#                             linelist.append(temp[1])
#                             
#                     LineTemp = fdata.readline()
#                     if len(LineTemp.strip())>0:
#                         LineTemp=LineTemp[0:-2] 
#                     LineTemp1=LineTemp.split("|");
#                     if LineTemp1[0]=="0" and LineTemp1[2]=="1":
#                         LineTemp = fdata.readline()
#                         if len(LineTemp.strip())>0:
#                             LineTemp=LineTemp[0:-2] 
#                         LineTemp1=LineTemp.split("|");
#                     
#                         if zjzh==undefined:
#                             if len(LineTemp1)>1:
#                                 
#                                 linelist.extend([LineTemp1[0]])
#                             if len(linelist)!=0:
#                                 listexcel.append([LineTemp1[0]])  
#                         else:
#                             if LineTemp1[0]==zjzh:    
# #                                 print LineTemp1[1]
#                                 if len(LineTemp1)>1:
#                                     linelist.extend([LineTemp1[0]])
#                                 if len(linelist)!=0:
#                                     listexcel.append(linelist)
#                     LineTemp = fdata.readline()
        
                        
                elif findstr=="Э��:202":
                    
                    linelist.append(file);
                    for index in range(len(list1)):
                        temp=list1[index].split(":")
                        if len(temp)==2:
                            linelist.append(temp[1])
                    LineTemp = fdata.readline()
                    if len(LineTemp.strip())>0:
                        LineTemp=LineTemp[0:-2] 
                    LineTemp1=LineTemp.split("|");
                    
                    if zjzh==undefined:
                        if len(LineTemp1)>1:
                            linelist.extend(LineTemp1)
                        if len(linelist)!=0:
                            listexcel.append(linelist)  
                    else:
                        print LineTemp1[2],zjzh
                        if LineTemp1[2]==zjzh:    
                            if len(LineTemp1)>1:
                                linelist.extend(LineTemp1)
                            if len(linelist)!=0:
                                listexcel.append(linelist)
                    
                    LineTemp = fdata.readline()
                else:
                    LineTemp = fdata.readline()
        else:        
            LineTemp = fdata.readline()
    fdata.close()
    return listexcel
   
    

if __name__=="__main__":
    global g_filepath,g_apitype,g_zjzh,g_excelfile
    filepath=g_filepath
    apitype=g_apitype
    zjzh=g_zjzh
    excelfilepath=g_excelfile
    
    if os.path.isdir(filepath):
        list = os.listdir(filepath)
        for i in range(0,len(list)):
            path = os.path.join(filepath,list[i])
            if os.path.isfile(path):
#                 print os.path.basename(path)
                filename=os.path.abspath(path) 
                filelist=jwkj_get_filePath_fileName_fileExt(filename)
#                 print filelist
                hzm=filelist[2]
                dates=filelist[1];
#                 print hzm
                if hzm==".log":
                    listexcel=ReadLog(filelist[0],filelist[1],filelist[2],apitype,zjzh)
                    if len(listexcel)>0:
                        if (apitype==100):
                            listexcel.insert(0,[u"����",u"ʱ��",u"�߳�",u"Э���",u"�ػ�ID",u"�ʽ��˺�",u"IP"
                                            ,u"Ӫҵ��",u"δ֪1",u"δ֪2",u"�ʽ��˺�",u"δ֪3",u"�ͻ���"])
#                         elif (apitype==101):
#                             listexcel.insert(0,[u"����",u"ʱ��",u"�߳�",u"Э���",u"�ػ�ID",u"�ͻ���"])
                        elif (apitype==202):
                            listexcel.insert(0,[u"����",u"ʱ��",u"�߳�",u"Э���",u"�ػ�ID",u"IP"
                                            ,u"Ӫҵ��",u"�ͻ���",u"δ֪1",u"δ֪2",u"δ֪3",u"�ɶ�����"
                                            ,u"δ֪4",u"δ֪5",u"��Ʊ����",u"ί�м۸�",u"δ֪6",u"ί������"
                                            ,u"δ֪7"])
                        RriteExcel(g_excelfile, listexcel)
                    print listexcel